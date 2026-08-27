import os
import re
import io
import uuid
import secrets
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from typing import Optional, List
from fastapi import FastAPI, HTTPException, status, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field
from supabase import create_client, Client
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Daily QR "din" isi timezone ke hisaab se decide hota hai (raat 12 baje PKT
# purana token khud-ba-khud expire ho jata hai — koi cron job nahi chahiye).
KARACHI_TZ = ZoneInfo("Asia/Karachi")

# Batch Attendance Register export ke liye: humara system session ki
# duration track nahi karta (sirf scan timestamp), is liye "Hours Attended"
# is fixed assumption se compute hota hai (reference file me bhi consistently
# 3 dikhi). Agar sessions ki length alag ho, ye constant update kar dein.
HOURS_PER_SESSION = 3


def today_in_karachi() -> str:
    """Aaj ki date Asia/Karachi timezone me, 'YYYY-MM-DD' format me."""
    return datetime.now(KARACHI_TZ).date().isoformat()


def today_start_utc_iso() -> str:
    """Aaj (Asia/Karachi) ke din ki shuruat, UTC isoformat me — same-day
    duplicate checks (device_id ya student ke against) isi boundary se
    'aaj' ka matlab decide karte hain."""
    return datetime.combine(
        datetime.now(KARACHI_TZ).date(), datetime.min.time(), tzinfo=KARACHI_TZ
    ).astimezone(timezone.utc).isoformat()


def parse_supabase_timestamp(iso_timestamp: str) -> datetime:
    """Supabase/Postgres timestamps ke fractional seconds kabhi kabhi
    non-standard length (5 digits jaise .11701, trailing zero trim ho
    jane se) me aate hain, jo Python 3.10 se pehle ka fromisoformat
    reject kar deta hai (sirf 3 ya 6-digit accept karta hai). Hamesha
    6-digit microseconds me normalize karke parse karte hain."""
    ts = iso_timestamp.replace("Z", "+00:00")
    ts = re.sub(r"\.(\d+)", lambda m: "." + m.group(1).ljust(6, "0")[:6], ts)
    return datetime.fromisoformat(ts)


def to_karachi_date(iso_timestamp: str):
    """Supabase ka timestamptz string -> Asia/Karachi local date (date object)."""
    return parse_supabase_timestamp(iso_timestamp).astimezone(KARACHI_TZ).date()


def roll_number_sort_key(roll_number: str):
    """UNITY-2 ko UNITY-10 se pehle rakhta hai (numeric suffix se sort,
    string-sort se nahi) — Admin Dashboard ke frontend sort ke consistent."""
    match = re.match(r"^(.*?)-(\d+)$", roll_number or "")
    if match:
        return (match.group(1), int(match.group(2)))
    return (roll_number or "", -1)


def _name_words(name: str) -> set:
    """Naam ko lowercase words me todta hai (punctuation ignore, 2 harf se
    chhoti words/initials jaise 'M.' ignore) — matching ke liye."""
    cleaned = re.sub(r"[^\w\s]", " ", (name or "").lower())
    return {w for w in cleaned.split() if len(w) >= 3}


def name_match_level(typed_name: str, registered_name: str) -> str:
    """
    Return 'exact' | 'partial' | 'none'.
    - 'exact': case-insensitive poora match (whitespace ke bawajood)
    - 'partial': kam se kam ek meaningful word common hai (typo/abbreviation
      jaisa tolerate karte hain, e.g. "M. Hamdan" vs "Muhammad Hamdan")
    - 'none': koi meaningful overlap nahi (jaisa "Safeer" vs "Muhammad Hamdan")
      — is case me attendance BLOCK hogi.
    """
    typed_clean = (typed_name or "").strip().lower()
    registered_clean = (registered_name or "").strip().lower()
    if typed_clean == registered_clean:
        return "exact"

    typed_words = _name_words(typed_name)
    registered_words = _name_words(registered_name)
    if typed_words and registered_words and (typed_words & registered_words):
        return "partial"

    return "none"

# Environment variables load karein
load_dotenv()

# Keys ab sirf .env file se aayengi (koi hardcoded secret source me nahi)
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_ANON_KEY") or os.getenv("SUPABASE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("WARNING: Supabase URL ya Key missing hain. Check environment variables.")

# Supabase Client Initialize Karein
try:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
except Exception as init_err:
    print(f"Supabase Client Init Failed: {init_err}")

app = FastAPI(
    title="Student Attendance Portal API",
    description="FastAPI Backend for QR-based Student Attendance System",
    version="1.0.0"
)

# FIXED CORS Configuration (All Origins, Methods & Headers Explicitly Allowed)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)


# --- Data Models (Pydantic) ---

class AttendanceCreateRequest(BaseModel):
    student_id: str = Field(..., json_schema_extra={"example": "CS-2026-045"}, description="Student Roll Number or ID")
    student_name: str = Field(..., json_schema_extra={"example": "Ali Ahmed"}, description="Student's Full Name")
    course_id: str = Field(..., json_schema_extra={"example": "8f3b2a1c-1234-5678-90ab-cdef12345678"}, description="Course UUID")
    token: str = Field(..., json_schema_extra={"example": "aBc123..."}, description="Today's QR code token")
    device_id: Optional[str] = Field(None, description="Anonymous per-browser device ID, used to block repeat scans from the same device on the same day")

class AttendanceResponse(BaseModel):
    status: str
    message: str
    record_id: Optional[str] = None
    captured_at: Optional[str] = None
    warning: Optional[str] = None


class QrTokenResponse(BaseModel):
    token: str
    valid_date: str
    course_id: Optional[str] = None


class ManualAttendanceRequest(BaseModel):
    course_id: str = Field(..., description="Course UUID")
    roll_number: str = Field(..., description="Student's roll number")


class ManualAttendanceResponse(BaseModel):
    status: str
    message: str
    record_id: Optional[str] = None
    already_marked: bool = False


# --- API Endpoints ---

@app.get("/")
def read_root():
    return {"status": "online", "message": "Attendance Portal Backend API working successfully!"}


@app.get("/health")
def health_check():
    """Railway (aur koi bhi uptime monitor) isi route se check karta hai
    ke app zinda hai — hamesha 200 return karta hai, DB tak touch nahi karta."""
    return {"status": "ok"}


@app.post("/api/v1/attendance", response_model=AttendanceResponse, status_code=status.HTTP_201_CREATED)
async def mark_attendance(payload: AttendanceCreateRequest):
    """
    React Frontend se student details aur image URL receive karke Supabase DB me save karta hai.
    """
    try:
        clean_student_id = payload.student_id.strip()
        clean_student_name = payload.student_name.strip()
        clean_course_id = payload.course_id.strip()
        clean_token = payload.token.strip()

        # --- Insert se pehle validation (garbage value DB tak na pahunche) ---
        if not clean_student_id or not clean_student_name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Student ID and Student Name cannot be empty."
            )

        try:
            uuid.UUID(clean_course_id)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"course_id ('{clean_course_id}') is not a valid UUID. "
                    "Please select the course again from the frontend dropdown and try again."
                )
            )

        # --- Verify the daily QR token ---
        if not clean_token:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No QR code scanned. Please scan today's QR code to mark attendance."
            )

        token_lookup = supabase.table("daily_qr_tokens").select("*").eq("token", clean_token).execute()

        if not token_lookup.data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid QR code. Please scan again and try."
            )

        token_record = token_lookup.data[0]
        if str(token_record.get("valid_date")) != today_in_karachi():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="QR code expired, please scan today's code."
            )

        token_course_id = token_record.get("course_id")
        if token_course_id and token_course_id != clean_course_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="This QR code belongs to a different course. Please scan the correct QR code."
            )

        # --- Same device can't mark attendance for this course twice in
        # one day (per-browser device_id, sent by the frontend). Blocks
        # one phone being used to check in multiple different students. ---
        clean_device_id = (payload.device_id or "").strip()
        if clean_device_id:
            device_dup = supabase.table("attendance_logs") \
                .select("id") \
                .eq("device_id", clean_device_id) \
                .eq("course_id", clean_course_id) \
                .gte("captured_at", today_start_utc_iso()) \
                .execute()
            if device_dup.data:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Attendance has already been marked from this device for this course today."
                )

        # --- Each course's students live in their own table — first look up
        # courses.students_table to find which table this course_id's students
        # are in. ---
        course_lookup = supabase.table("courses") \
            .select("students_table") \
            .eq("id", clean_course_id) \
            .execute()

        students_table = course_lookup.data[0].get("students_table") if course_lookup.data else None
        if not students_table:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Students table is not configured for this course. Please contact the admin."
            )

        # --- Sirf REGISTERED students hi attendance mark kar sakein ---
        # Case-insensitive match (ilike): purane batches uppercase roll
        # numbers use karte the (UNITY-1), naye Batch 4 lowercase hain
        # (unityk4-01) — dono ko student phone par chhota/bada harf type
        # kare to bhi match hona chahiye.
        student_lookup = supabase.table(students_table) \
            .select("id, roll_number, name") \
            .ilike("roll_number", clean_student_id) \
            .execute()

        if not student_lookup.data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="This Student ID is not registered. Please get registered by the admin."
            )

        registered_student = student_lookup.data[0]
        registered_name = registered_student.get("name")
        name_warning = None

        if not registered_name:
            # Batch 4 jaise students bulk-register hote waqt naam pata nahi
            # tha (sirf roll_number diya gaya tha). Ye pehli attendance hai —
            # jo naam ab submit hua hai, wahi ab se canonical ban jata hai.
            update_resp = supabase.table(students_table) \
                .update({"name": clean_student_name}) \
                .eq("id", registered_student["id"]) \
                .execute()
            if update_resp.data:
                registered_student = update_resp.data[0]
            else:
                # RLS ya kisi wajah se update fail ho jaye to bhi ye attendance
                # NOT NULL student_name ki wajah se crash na ho — submitted
                # naam hi use kar lo is record ke liye.
                registered_student = {**registered_student, "name": clean_student_name}
        else:
            # Naam check — teen levels:
            #  exact   -> bilkul theek, koi warning nahi
            #  partial -> kam se kam ek meaningful word match karta hai (typo/
            #             abbreviation jaisa, e.g. "M. Hamdan" vs "Muhammad Hamdan")
            #             -> allow, lekin warning ke sath
            #  none    -> koi overlap hi nahi (jaisa "Safeer" vs "Muhammad Hamdan")
            #             -> BLOCK, kisi aur ke roll number par galat naam se
            #             attendance lagne se roko
            match_level = name_match_level(clean_student_name, registered_name)

            if match_level == "none":
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        f"This Student ID '{registered_student['roll_number']}' is registered "
                        f"under a different student's name. Please use your correct Student ID, "
                        "or get registered by the admin."
                    )
                )

            if match_level == "partial":
                name_warning = (
                    f"Note: the name you entered ('{clean_student_name}') does not fully match "
                    f"the registered record ('{registered_name}'). "
                    "Attendance was still marked under the registered name."
                )
            # registered_name (existing) ko canonical rakho, submitted naya
            # naam se overwrite mat karo — ye sirf naam SET karne ke liye hai,
            # UPDATE karne ke liye nahi.

        # 1. Server-side timezone-aware UTC timestamp generate karein
        current_time = datetime.now(timezone.utc).isoformat()

        # 2. Directly 'attendance_logs' table me insert karein — student_id/
        # student_name yahan form ki raw value se nahi, balke students table
        # ke CANONICAL record se aate hain (typo-proof), aur registered_student_id
        # asal FK (uuid) students.id ki taraf point karta hai.
        record_data = {
            "student_id": registered_student["roll_number"],
            "student_name": registered_student["name"],
            "registered_student_id": registered_student["id"],
            "course_id": clean_course_id,
            "captured_at": current_time,
            "device_id": clean_device_id or None,
        }

        # Debug log: agla error aaye to terminal me exact payload turant dikhega
        print(f"[mark_attendance] Inserting record: {record_data}")

        response = supabase.table("attendance_logs").insert(record_data).execute()

        if not response.data or len(response.data) == 0:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Could not insert entry into the database."
            )

        inserted_record = response.data[0]

        return AttendanceResponse(
            status="success",
            message="Attendance successfully done!",
            record_id=str(inserted_record.get("id", "")),
            captured_at=str(inserted_record.get("captured_at", current_time)),
            warning=name_warning
        )

    except HTTPException as http_err:
        raise http_err
    except Exception as err:
        err_str = str(err)
        print(f"Error saving attendance: {err_str}")
        
        if "401" in err_str or "Invalid API key" in err_str:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Supabase API key is invalid or expired. Check the server's .env / SUPABASE_KEY."
            )
            
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database / Server Error: {err_str}"
        )


@app.post("/api/v1/attendance/manual", response_model=ManualAttendanceResponse, status_code=status.HTTP_201_CREATED)
async def mark_attendance_manual(payload: ManualAttendanceRequest):
    """
    Admin Dashboard se seedha attendance mark karta hai — QR token ya
    photo ki zaroorat nahi (admin khud "Registered Students" list se
    student select karta hai). Same-day duplicate insert nahi karta.
    """
    try:
        clean_course_id = payload.course_id.strip()
        clean_roll = payload.roll_number.strip()

        try:
            uuid.UUID(clean_course_id)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"course_id ('{clean_course_id}') is not a valid UUID."
            )

        course_lookup = supabase.table("courses").select("students_table").eq("id", clean_course_id).execute()
        students_table = course_lookup.data[0].get("students_table") if course_lookup.data else None
        if not students_table:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Students table is not configured for this course."
            )

        student_lookup = supabase.table(students_table) \
            .select("id, roll_number, name") \
            .ilike("roll_number", clean_roll) \
            .execute()

        if not student_lookup.data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="This Student ID is not registered."
            )

        student = student_lookup.data[0]
        if not student.get("name"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "This student's name is not set yet — the student must scan once "
                    "themselves first to give their first attendance (so the name gets saved), "
                    "only then can manual marking be used."
                )
            )

        # Aaj (Asia/Karachi) ke liye duplicate check — dobara insert na ho
        existing = supabase.table("attendance_logs") \
            .select("id") \
            .eq("registered_student_id", student["id"]) \
            .eq("course_id", clean_course_id) \
            .gte("captured_at", today_start_utc_iso()) \
            .execute()

        if existing.data:
            return ManualAttendanceResponse(
                status="success",
                message=f"{student['name']}'s attendance is already marked for today.",
                record_id=existing.data[0]["id"],
                already_marked=True,
            )

        current_time = datetime.now(timezone.utc).isoformat()
        record_data = {
            "student_id": student["roll_number"],
            "student_name": student["name"],
            "registered_student_id": student["id"],
            "course_id": clean_course_id,
            "captured_at": current_time,
        }

        response = supabase.table("attendance_logs").insert(record_data).execute()
        if not response.data:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Could not insert entry into the database."
            )

        return ManualAttendanceResponse(
            status="success",
            message=f"{student['name']}'s attendance has been marked.",
            record_id=str(response.data[0].get("id", "")),
        )

    except HTTPException as http_err:
        raise http_err
    except Exception as err:
        print(f"Error in manual attendance: {err}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database / Server Error: {err}"
        )


@app.get("/api/v1/qr/today", response_model=QrTokenResponse)
async def get_today_qr_token(course_id: Optional[str] = None):
    """
    Aaj ka valid daily-QR token laao. Agar aaj ke liye already generate ho
    chuka hai to wahi return hota hai (naya nahi banta). Naya din shuru hone
    par (Asia/Karachi midnight) khud-ba-khud naya token ban jata hai.
    course_id diya jaye to per-course QR (abhi optional/future-ready), warna
    global daily QR (course_id = NULL) use hota hai.
    """
    try:
        clean_course_id = None
        if course_id:
            clean_course_id = course_id.strip()
            try:
                uuid.UUID(clean_course_id)
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"course_id ('{clean_course_id}') is not a valid UUID."
                )

        today = today_in_karachi()

        query = supabase.table("daily_qr_tokens").select("*").eq("valid_date", today)
        query = query.is_("course_id", "null") if clean_course_id is None else query.eq("course_id", clean_course_id)
        existing = query.execute()

        if existing.data:
            row = existing.data[0]
            return QrTokenResponse(token=row["token"], valid_date=str(row["valid_date"]), course_id=row.get("course_id"))

        # Naya token generate karein (crypto-secure random, date se derive nahi)
        new_token = secrets.token_urlsafe(24)
        insert_data = {"token": new_token, "valid_date": today, "course_id": clean_course_id}

        try:
            inserted = supabase.table("daily_qr_tokens").insert(insert_data).execute()
            row = inserted.data[0]
        except Exception as insert_err:
            # Race condition: dusri request ne isi waqt insert kar diya (unique index).
            # Existing row dobara fetch kar lein.
            print(f"[qr/today] Insert race, re-fetching: {insert_err}")
            retry = query.execute()
            if not retry.data:
                raise
            row = retry.data[0]

        return QrTokenResponse(token=row["token"], valid_date=str(row["valid_date"]), course_id=row.get("course_id"))

    except HTTPException as http_err:
        raise http_err
    except Exception as err:
        print(f"Error generating QR token: {err}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Could not generate QR token: {err}"
        )


@app.get("/api/v1/export/attendance-report")
async def export_attendance_report():
    """
    CEGA "Batch Attendance Register" format me poora attendance data
    .xlsx export karta hai — har course ki apni sheet, har registered
    student ki row (chahe kabhi scan hua ho ya na ho), har training-day
    (jis din us course me kam se kam ek scan hua) ka P/A column.
    """
    try:
        courses = (supabase.table("courses").select("id, code, name, students_table").execute().data) or []
        attendance = (
            supabase.table("attendance_logs")
            .select("registered_student_id, course_id, captured_at")
            .execute()
            .data
        ) or []

        # course_id -> training-day dates (kisi bhi scan se, legacy/unregistered records bhi count)
        course_dates = {}
        # (registered_student_id, course_id) -> present dates
        student_presence = {}

        for row in attendance:
            course_id = row.get("course_id")
            if not course_id or not row.get("captured_at"):
                continue
            try:
                day = to_karachi_date(row["captured_at"])
            except Exception:
                continue

            course_dates.setdefault(course_id, set()).add(day)

            reg_id = row.get("registered_student_id")
            if reg_id:
                student_presence.setdefault((reg_id, course_id), set()).add(day)

        # Har course apni alag table me hai — students_table ke hisaab se
        # fetch karein aur course_id ke against jama karein.
        students_by_course = {}
        for course in courses:
            table_name = course.get("students_table")
            if not table_name:
                continue
            rows = (supabase.table(table_name).select("id, roll_number, name").execute().data) or []
            rows.sort(key=lambda s: roll_number_sort_key(s["roll_number"]))
            students_by_course[course["id"]] = rows

        wb = Workbook()
        wb.remove(wb.active)

        title_font = Font(bold=True, size=13, color="FFFFFF")
        title_fill = PatternFill("solid", fgColor="2563EB")
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill("solid", fgColor="EFF6FF")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        used_sheet_names = set()

        for course in sorted(courses, key=lambda c: c["code"]):
            course_id = course["id"]
            dates = sorted(course_dates.get(course_id, set()))
            course_students = students_by_course.get(course_id, [])
            n_dates = len(dates)

            sheet_name = re.sub(r'[\\/*?:\[\]]', '-', course["code"])[:31] or "Course"
            base_name, n = sheet_name, 1
            while sheet_name in used_sheet_names:
                n += 1
                suffix = f"_{n}"
                sheet_name = base_name[: 31 - len(suffix)] + suffix
            used_sheet_names.add(sheet_name)

            ws = wb.create_sheet(title=sheet_name)

            hours_col = 3 + n_dates
            pct_col = hours_col + 1
            total_cols = max(pct_col, 3)

            # Row 1 — title
            display_name = (course.get("name") or course["code"]).strip().upper()
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            title_cell = ws.cell(row=1, column=1, value=f"CEGA — {display_name}, ATTENDANCE REGISTER")
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = center
            ws.row_dimensions[1].height = 22

            # Rows 2-3 — headers (S.No / Name / Hrs-Day / dates / Hours Attended / Attendance %)
            for col in (1, 2, hours_col, pct_col):
                ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

            ws.cell(row=2, column=1, value="S.No")
            ws.cell(row=2, column=2, value="Student Name")
            ws.cell(row=2, column=hours_col, value="Hours\nAttended")
            ws.cell(row=2, column=pct_col, value="Attendance\n%")

            if n_dates > 0:
                ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=2 + n_dates)
                ws.cell(row=2, column=3, value=f"Hrs/Day → {HOURS_PER_SESSION}")

            for col in range(1, total_cols + 1):
                cell = ws.cell(row=2, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border

            for i, day in enumerate(dates):
                cell = ws.cell(row=3, column=3 + i, value=day.strftime("%d-%b\n%a"))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border

            # Student rows
            row_idx = 4
            for i, student in enumerate(course_students, start=1):
                sno_cell = ws.cell(row=row_idx, column=1, value=i)
                sno_cell.alignment = center
                sno_cell.border = border
                name_cell = ws.cell(row=row_idx, column=2, value=student["name"])
                name_cell.border = border

                present_dates = student_presence.get((student["id"], course_id), set())
                p_count = 0
                for j, day in enumerate(dates):
                    is_present = day in present_dates
                    if is_present:
                        p_count += 1
                    cell = ws.cell(row=row_idx, column=3 + j, value="P" if is_present else "A")
                    cell.alignment = center
                    cell.border = border
                    cell.font = Font(bold=True, color="15803D" if is_present else "B91C1C")

                hours_cell = ws.cell(row=row_idx, column=hours_col, value=p_count * HOURS_PER_SESSION)
                hours_cell.alignment = center
                hours_cell.border = border

                pct = round((p_count / n_dates) * 100, 1) if n_dates > 0 else 0
                pct_cell = ws.cell(row=row_idx, column=pct_col, value=f"{pct}%")
                pct_cell.alignment = center
                pct_cell.border = border

                row_idx += 1

            if not course_students:
                ws.cell(row=4, column=1, value="(No students registered for this course)")

            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 26
            for i in range(n_dates):
                ws.column_dimensions[get_column_letter(3 + i)].width = 9
            ws.column_dimensions[get_column_letter(hours_col)].width = 10
            ws.column_dimensions[get_column_letter(pct_col)].width = 10
            ws.freeze_panes = ws.cell(row=4, column=3)

        if not wb.sheetnames:
            wb.create_sheet(title="No Data")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"Batch_Attendance_Register_{today_in_karachi()}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except Exception as err:
        print(f"Error generating attendance report: {err}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Could not generate attendance report: {err}"
        )


@app.get("/api/v1/attendance/{course_id}")
async def get_course_attendance(course_id: str):
    """
    Teacher ya Admin dashboard ke liye specific course ki tamam attendance logs get karein.
    """
    try:
        response = supabase.table("attendance_logs")\
            .select("id, student_id, student_name, captured_at")\
            .eq("course_id", course_id)\
            .order("captured_at", desc=True)\
            .execute()

        return {
            "status": "success",
            "course_id": course_id,
            "total_records": len(response.data),
            "logs": response.data
        }
    except Exception as err:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(err))

# Dynamic Port Execution for Railway
if __name__ == "__main__":
    import uvicorn
    # Default port ko 8080 kar diya hai taake Railway ke sath perfect match ho
    port = int(os.environ.get("PORT", 8080))
    uvicorn.run("main:app", host="0.0.0.0", port=port)